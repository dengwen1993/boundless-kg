"""Lightweight file parsing for tmp uploads — classification + short preview.

Designed for 「简单解析分类」 — not high-fidelity content extraction.
Output goal: enough text for the LLM to know what the file is about,
who wrote it, and the first ~2KB of body content.  Anything more
sophisticated should defer to the bundled skills:

  * PPTX → src/skills/pptx-generator (uses ``markitdown`` when available)
  * DOCX → src/skills/minimax-docx (uses ``pandoc`` → falls back to XML)

This module deliberately avoids adding heavy deps (python-docx,
python-pptx, markitdown) — the only formats that need an external lib
are PDF (``pypdf``) and XLSX (``openpyxl``), both already in
``pyproject.toml``.  Everything else uses the stdlib.

Format coverage:

* text-like    → UTF-8 / gb18030 / latin-1 fallback
* CSV          → stdlib ``csv`` → markdown table (first 200 rows)
* HTML         → stdlib regex strip (no ``bs4`` dep)
* PDF          → ``pypdf``
* DOCX         → stdlib ``zipfile`` + XML strip (same as ``docx_preview.sh``
                  fallback path); classification hints from core.xml
* PPTX         → stdlib ``zipfile`` + XML strip across ``ppt/slides/*.xml``
* XLSX         → ``openpyxl`` (already required for graph store writes)
* images       → ``PIL`` metadata only (no pixel / no OCR)
"""

from __future__ import annotations

import csv
import io
import logging
import re
import zipfile
from html import unescape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)


# Bytes-per-char heuristic when streaming the result back to the agent.
# UTF-8 averages ~3 bytes/char for CJK, ~1.5 for Latin; we use 2 as a
# safe upper bound so we don't accidentally truncate a doc that's
# mostly ASCII just to hit an exact char count.
_BYTE_PER_CHAR: float = 2.0

#: Extensions we can read directly with stdlib alone (no extra deps).
_PLAIN_TEXT_EXTS: set[str] = {
    ".txt", ".md", ".rst", ".log",
    ".json", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf",
    ".py", ".js", ".ts", ".tsx", ".jsx", ".mjs", ".cjs",
    ".go", ".rs", ".java", ".kt", ".scala",
    ".c", ".cpp", ".h", ".hpp", ".cc", ".cxx",
    ".sh", ".bash", ".zsh", ".fish",
    ".sql", ".graphql", ".proto",
    ".xml", ".tex", ".bib",
}

_IMAGE_EXTS: set[str] = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}


async def parse_file_to_text(path: Path, *, max_chars: int = 8000) -> dict[str, Any]:
    """Extract a short preview + classification hints from *path*.

    Returns ``{"text": ..., "truncated": bool, "format": ..., "size": ...,
    "hints": {...}}`` — ``hints`` carries structured metadata
    (slide count, sheet names, author, etc.) so the LLM can answer
    「这份 PPT 多少页？」 without re-reading the body.

    Never raises on unknown formats — instead returns a clear "format
    not supported" string in ``text`` so the agent can recover
    gracefully instead of crashing the tool call.
    """
    if not path.exists():
        raise ValueError(f"文件不存在：{path}")
    if not path.is_file():
        raise ValueError(f"不是文件：{path}")

    suffix = path.suffix.lower()
    size = path.stat().st_size
    byte_cap = int(max_chars * _BYTE_PER_CHAR)

    try:
        if suffix in _PLAIN_TEXT_EXTS or suffix == "":
            text, hints = _read_text(path, byte_cap), {"kind": "text"}
            fmt = "text"
        elif suffix == ".csv":
            text, hints = _read_csv(path, byte_cap), {"kind": "tabular"}
            fmt = "csv"
        elif suffix in {".html", ".htm"}:
            text, hints = _read_html(path, byte_cap), {"kind": "html"}
            fmt = "html"
        elif suffix == ".pdf":
            text, hints = _read_pdf(path, byte_cap)
            fmt = "pdf"
        elif suffix == ".docx":
            text, hints = _read_docx(path, byte_cap)
            fmt = "docx"
        elif suffix == ".pptx":
            text, hints = _read_pptx(path, byte_cap)
            fmt = "pptx"
        elif suffix in {".xlsx", ".xlsm"}:
            text, hints = _read_xlsx(path, byte_cap)
            fmt = "xlsx"
        elif suffix in _IMAGE_EXTS:
            text, hints = _read_image_meta(path), {"kind": "image"}
            fmt = "image"
        else:
            text = (
                f"⚠️ 暂不支持直接解析 ``{suffix or '<no-ext>'}`` 格式的文件。"
                f"如需解析请先把文件转成纯文本 / Markdown 再上传。"
            )
            hints = {"kind": "unknown"}
            fmt = "unsupported"
    except Exception as exc:
        logger.exception("parse failed for %s", path.name)
        text = (
            f"❌ 解析失败（{type(exc).__name__}: {exc}）。"
            f"文件可能已损坏、加密或为不支持的格式变体。"
        )
        hints = {"kind": "error"}
        fmt = "error"

    truncated = len(text) >= max_chars
    return {
        "text": text,
        "truncated": truncated,
        "format": fmt,
        "size": size,
        "chars": len(text),
        "hints": hints,
    }


# ------------------------------------------------------------------
# Format-specific readers — all return (text, hints)
# ------------------------------------------------------------------


def _read_text(path: Path, byte_cap: int) -> str:
    """Read a plain-text file with UTF-8 / latin-1 fallback."""
    raw = path.read_bytes()[:byte_cap]
    for enc in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _read_csv(path: Path, byte_cap: int) -> tuple[str, dict[str, Any]]:
    """Render a CSV as a markdown table (first 200 rows)."""
    raw = path.read_text(encoding="utf-8", errors="replace")
    if len(raw.encode("utf-8")) > byte_cap:
        raw = raw.encode("utf-8")[:byte_cap].decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(raw))
    rows: list[list[str]] = []
    max_rows = 200
    max_cell_width = 80
    total_rows = 0
    for i, row in enumerate(reader):
        total_rows = i + 1
        if i >= max_rows:
            rows.append(["…", "（后续行已截断）", ""])
            break
        rows.append([
            cell[:max_cell_width] + ("…" if len(cell) > max_cell_width else "")
            for cell in row
        ])
    if not rows:
        return "（空 CSV 文件）", {"rows": 0, "cols": 0}

    widths = [max(len(r[i]) if i < len(r) else 0 for r in rows) for i in range(max(len(r) for r in rows))]

    def fmt_row(r: list[str]) -> str:
        return "| " + " | ".join((r[i] if i < len(r) else "").ljust(widths[i]) for i in range(len(widths))) + " |"

    sep = "| " + " | ".join("-" * w for w in widths) + " |"
    body = "\n".join([fmt_row(rows[0]), *[fmt_row(r) for r in rows[1:]]])
    text = body + (f"\n\n（共 {total_rows}+ 行，仅显示前 {max_rows} 行）" if total_rows > max_rows else "")
    return text, {"rows": total_rows, "cols": len(rows[0]) if rows else 0, "kind": "tabular"}


def _read_html(path: Path, byte_cap: int) -> str:
    """Strip HTML tags and return visible text — stdlib only, no bs4."""
    raw = path.read_text(encoding="utf-8", errors="replace")
    if len(raw.encode("utf-8")) > byte_cap:
        raw = raw.encode("utf-8")[:byte_cap].decode("utf-8", errors="replace")
    # Drop script / style content BEFORE stripping tags — otherwise
    # inline JS / CSS shows up as garbage text in the preview.
    raw = re.sub(r"<script\b[^>]*>.*?</script>", "", raw, flags=re.DOTALL | re.IGNORECASE)
    raw = re.sub(r"<style\b[^>]*>.*?</style>", "", raw, flags=re.DOTALL | re.IGNORECASE)
    raw = re.sub(r"<!--.*?-->", "", raw, flags=re.DOTALL)
    # Replace block-level closers with newlines so the output keeps some structure.
    raw = re.sub(r"</(p|div|br|h[1-6]|li|tr)\s*>", "\n", raw, flags=re.IGNORECASE)
    raw = re.sub(r"<[^>]+>", "", raw)
    return unescape(raw)


def _read_pdf(path: Path, byte_cap: int) -> tuple[str, dict[str, Any]]:
    """Extract text from a PDF using ``pypdf``."""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    page_count = len(reader.pages)
    parts: list[str] = [f"📕 PDF · {page_count} 页\n"]
    char_count = 0
    for page_idx, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception as exc:  # pragma: no cover — depends on PDF
            page_text = f"\n[第 {page_idx + 1} 页解析失败: {exc}]\n"
        parts.append(f"\n--- 第 {page_idx + 1} 页 ---\n{page_text}")
        char_count += len(page_text)
        if char_count > byte_cap:
            parts.append("\n…（后续页已截断）")
            break
    hints = {"kind": "pdf", "pages": page_count}
    # Try to read metadata (author / title) — best-effort, swallow errors.
    try:
        meta = reader.metadata or {}
        if meta:
            hints["author"] = str(meta.get("/Author", "")).strip() or None
            hints["title"] = str(meta.get("/Title", "")).strip() or None
    except Exception:
        pass
    return "".join(parts) or "（PDF 没有可提取的文本层，可能是扫描件）", hints


# ------------------------------------------------------------------
# DOCX / PPTX — markitdown first (best quality), stdlib zipfile as fallback
# ------------------------------------------------------------------

_markitdown_instance: Any | None = None
_markitdown_available: bool | None = None


def _get_markitdown() -> Any:
    """Lazily build a single :class:`MarkItDown` instance.

    Returns ``None`` when markitdown isn't installed so the callers
    can fall through to the stdlib zipfile path.  We cache the result
    so the (non-trivial) magika model isn't re-loaded on every call.
    """
    global _markitdown_instance, _markitdown_available
    if _markitdown_available is False:
        return None
    if _markitdown_instance is not None:
        return _markitdown_instance
    try:
        from markitdown import MarkItDown  # type: ignore[import-not-found]

        _markitdown_instance = MarkItDown()
        _markitdown_available = True
        return _markitdown_instance
    except ImportError:
        _markitdown_available = False
        return None


def _parse_via_markitdown(path: Path) -> str | None:
    """Best-effort Office → Markdown via ``markitdown``.

    Returns ``None`` if markitdown isn't installed or fails (so the
    caller can fall back to the stdlib path).  ``markitdown`` raises
    on malformed Office files — we want to recover, not crash.
    """
    md = _get_markitdown()
    if md is None:
        return None
    try:
        result = md.convert(str(path))
        return result.text_content or None
    except Exception as exc:
        logger.warning("markitdown failed for %s, falling back: %s", path.name, exc)
        return None


# ------------------------------------------------------------------
# DOCX / PPTX — zero-dep readers using stdlib zipfile + XML strip
# Used as the fallback when markitdown is missing or fails on a
# malformed Office file.  Mirrors the fallback in
# src/skills/minimax-docx/scripts/docx_preview.sh.
# ------------------------------------------------------------------


def _strip_xml_to_text(xml_bytes: bytes) -> str:
    """Parse XML and concatenate all visible text nodes.

    Uses ``xml.etree.ElementTree`` (stdlib) — fast, safe against
    malformed XML (we already know it's a real zip from a real Office
    file, so the XML is well-formed in practice).
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        # Last-ditch: regex-strip tags.  Loses structure but keeps text.
        text = xml_bytes.decode("utf-8", errors="replace")
        return re.sub(r"<[^>]+>", "", text)
    # ET.iter guarantees document order, which is exactly what we want
    # — paragraph / run boundaries map to text-node order.
    parts: list[str] = []
    for node in root.iter():
        if node.text:
            parts.append(node.text)
        if node.tail:
            parts.append(node.tail)
    return " ".join(parts)


def _read_docx(path: Path, byte_cap: int) -> tuple[str, dict[str, Any]]:
    """Read DOCX → Markdown via markitdown (preferred), or stdlib zipfile.

    Prefers :func:`_parse_via_markitdown` because it preserves headings,
    lists, tables, and links as proper Markdown.  Falls back to the
    stdlib zipfile + XML strip path when markitdown is missing or fails
    (e.g. on a malformed DOCX).
    """
    hints: dict[str, Any] = {"kind": "docx"}

    # Best path: markitdown → clean Markdown.
    md_text = _parse_via_markitdown(path)
    if md_text:
        hints["engine"] = "markitdown"
        # Pull author / title / pages from the docx zip metadata even
        # when markitdown succeeds — those don't show up in the body.
        try:
            with zipfile.ZipFile(path) as zf:
                if "docProps/core.xml" in zf.namelist():
                    core_text = _strip_xml_to_text(zf.read("docProps/core.xml"))
                    # core.xml tags are concatenated without separators;
                    # squeeze whitespace so the hint is one short line.
                    hints["meta"] = " ".join(core_text.split())[:300]
                if "docProps/app.xml" in zf.namelist():
                    try:
                        app_meta = ET.fromstring(zf.read("docProps/app.xml"))
                        pages_el = app_meta.find(
                            ".//{http://schemas.openxmlformats.org/officeDocument/2006/extended-properties}Pages"
                        )
                        if pages_el is not None and pages_el.text:
                            hints["pages"] = int(pages_el.text)
                    except Exception:
                        pass
        except zipfile.BadZipFile:
            pass
        truncated = len(md_text) > byte_cap
        if truncated:
            md_text = md_text[:byte_cap] + "\n…（已截断）"
        return md_text, hints

    # Fallback: stdlib zipfile + XML strip.
    hints["engine"] = "stdlib"
    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
            if "word/document.xml" not in names:
                return "（DOCX 缺少 word/document.xml，可能已损坏）", hints

            doc_xml = zf.read("word/document.xml")
            body = _strip_xml_to_text(doc_xml)

            # Metadata hints — title / author / page count.
            if "docProps/core.xml" in names:
                try:
                    meta = _strip_xml_to_text(zf.read("docProps/core.xml"))
                    hints["meta"] = meta[:300]
                except Exception:
                    pass
            if "docProps/app.xml" in names:
                try:
                    app_meta = ET.fromstring(zf.read("docProps/app.xml"))
                    pages_el = app_meta.find(".//{http://schemas.openxmlformats.org/officeDocument/2006/extended-properties}Pages")
                    if pages_el is not None and pages_el.text:
                        hints["pages"] = int(pages_el.text)
                except Exception:
                    pass
    except zipfile.BadZipFile:
        return "❌ 文件不是有效的 DOCX（zip 解析失败）。", hints

    body = body.strip()
    if not body:
        return "（DOCX 没有可提取的文本，可能全是图片或表格）", hints

    truncated = len(body) > byte_cap
    if truncated:
        body = body[:byte_cap] + "\n…（已截断）"
    return body, hints


def _read_pptx(path: Path, byte_cap: int) -> tuple[str, dict[str, Any]]:
    """Read PPTX → Markdown via markitdown (preferred), or stdlib zipfile.

    Prefers :func:`_parse_via_markitdown` because it preserves slide
    structure, titles, and any embedded tables as proper Markdown.
    Falls back to reading each slide's XML directly when markitdown is
    missing or fails.
    """
    hints: dict[str, Any] = {"kind": "pptx"}

    # Best path: markitdown → Markdown with slide annotations.
    md_text = _parse_via_markitdown(path)
    if md_text:
        hints["engine"] = "markitdown"
        # markitdown inserts ``<!-- Slide number: N -->`` markers — count
        # them so the hint reflects the actual slide count.
        slide_count = md_text.count("<!-- Slide number:")
        if slide_count:
            hints["slides"] = slide_count
        truncated = len(md_text) > byte_cap
        if truncated:
            md_text = md_text[:byte_cap] + "\n…（已截断）"
        return md_text, hints

    # Fallback: stdlib zipfile + XML strip.
    hints["engine"] = "stdlib"
    try:
        with zipfile.ZipFile(path) as zf:
            slide_names = sorted(
                n for n in zf.namelist()
                if re.match(r"ppt/slides/slide\d+\.xml$", n)
            )
            if not slide_names:
                return "（PPTX 找不到 slide XML，可能已损坏）", hints
            hints["slides"] = len(slide_names)

            # Try core.xml for title / author.
            if "docProps/core.xml" in zf.namelist():
                try:
                    hints["meta"] = _strip_xml_to_text(zf.read("docProps/core.xml"))[:300]
                except Exception:
                    pass

            parts: list[str] = [f"📊 PPTX · {len(slide_names)} 页\n"]
            char_count = 0
            for idx, name in enumerate(slide_names, start=1):
                slide_text = _strip_xml_to_text(zf.read(name))
                parts.append(f"\n--- Slide {idx} ---\n{slide_text}")
                char_count += len(slide_text)
                if char_count > byte_cap:
                    parts.append("\n…（后续页已截断）")
                    break
            return "\n".join(parts), hints
    except zipfile.BadZipFile:
        return "❌ 文件不是有效的 PPTX（zip 解析失败）。", hints


def _read_xlsx(path: Path, byte_cap: int) -> tuple[str, dict[str, Any]]:
    """Render an XLSX as markdown tables — ``openpyxl`` (already a dep)."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    hints: dict[str, Any] = {
        "kind": "xlsx",
        "sheets": list(wb.sheetnames),
    }
    parts: list[str] = []
    char_count = 0
    max_rows = 100
    max_cell_width = 60

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n## Sheet: {sheet_name}\n")
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(["…", "（后续行已截断）", ""])
                break
            rows.append([
                (str(c) if c is not None else "")[:max_cell_width]
                + ("…" if c is not None and len(str(c)) > max_cell_width else "")
                for c in row
            ])
        if not rows:
            parts.append("（空 sheet）")
            continue
        widths = [
            max((len(r[i]) if i < len(r) else 0) for r in rows)
            for i in range(max(len(r) for r in rows))
        ]

        def fmt_row(r: list[str]) -> str:
            return "| " + " | ".join((r[i] if i < len(r) else "").ljust(widths[i]) for i in range(len(widths))) + " |"

        parts.append(fmt_row(rows[0]))
        parts.append("| " + " | ".join("-" * w for w in widths) + " |")
        for r in rows[1:]:
            parts.append(fmt_row(r))

        char_count += sum(len(s) for s in parts[-len(rows) - 2 :])
        if char_count > byte_cap:
            parts.append("\n…（后续 sheet 已截断）")
            break

    wb.close()
    text = "\n".join(parts) or "（XLSX 没有可提取的内容）"
    return text, hints


def _read_image_meta(path: Path) -> str:
    """Read image metadata via PIL — not pixel content.

    Pixel content would need a multimodal LLM call; for now we surface
    the metadata (size, mode, format, EXIF if present) so the agent at
    least knows what the file is.
    """
    from PIL import Image

    try:
        with Image.open(path) as img:
            width, height = img.size
            mode = img.mode
            fmt = img.format
            exif_summary: list[str] = []
            try:
                exif = img.getexif()
                if exif:
                    for tag_id, label in (
                        (0x010E, "ImageDescription"),
                        (0x010F, "Make"),
                        (0x0110, "Model"),
                        (0x9003, "DateTimeOriginal"),
                    ):
                        if tag_id in exif:
                            exif_summary.append(f"  - {label}: {exif[tag_id]}")
            except Exception:
                pass

        lines = [
            f"📷 图片文件：{path.name}",
            f"   格式：{fmt}，尺寸：{width}×{height}，模式：{mode}",
        ]
        if exif_summary:
            lines.append("   EXIF：")
            lines.extend(exif_summary)
        lines.append(
            "\n（当前仅提取元数据；如需 OCR 或图像内容理解，需要接入多模态 LLM。）"
        )
        return "\n".join(lines)
    except Exception as exc:
        return f"⚠️ 无法读取图片元数据：{exc}"


__all__ = ["parse_file_to_text"]